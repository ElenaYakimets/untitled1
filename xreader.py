from openpyxl import load_workbook

filename = "data.xlsx"

if __name__ == "__main__":
    wb = load_workbook(filename)
    print(wb.sheetnames)