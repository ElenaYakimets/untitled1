from openpyxl import load_workbook, Workbook

filename = "data.xlsx"

if __name__ == "__main__":
    wb: Workbook = load_workbook(filename)
    sheet = wb.active
    c1 = sheet.cell(row=1, column=2)
    c1.value = 'B269OT177/BM032077'
    c1 = sheet.cell(row=2, column=2)
    c1.value = '10013160/130920/0493176'
        # sheet.title = 'Sheet1'
        # print('sheet name is renamed as:'+ sheet.title)
    wb.save(filename)